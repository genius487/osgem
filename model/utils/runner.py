# Defines the modules

import logging
import os

import pandas as pd
from utils.onsset import (SET_ELEC_ORDER, SET_LCOE_GRID, SET_MIN_GRID_DIST, SET_GRID_PENALTY,
                    SET_MV_CONNECT_DIST, SET_WINDVEL, SET_WINDCF, SettlementProcessor, Technology)

try:
    from onsset.specs import (SPE_COUNTRY, SPE_ELEC, SPE_ELEC_MODELLED,
                              SPE_ELEC_RURAL, SPE_ELEC_URBAN, SPE_END_YEAR, SPE_MID_YEAR,
                              SPE_GRID_CAPACITY_INVESTMENT, SPE_GRID_LOSSES,
                              SPE_MAX_GRID_EXTENSION_DIST,
                              SPE_NUM_PEOPLE_PER_HH_RURAL,
                              SPE_NUM_PEOPLE_PER_HH_URBAN, SPE_POP, SPE_POP_FUTURE,
                              SPE_START_YEAR, SPE_URBAN, SPE_URBAN_FUTURE,
                              SPE_URBAN_MODELLED)
except ImportError:
    from utils.specs import (SPE_COUNTRY, SPE_ELEC, SPE_ELEC_MODELLED,
                       SPE_ELEC_RURAL, SPE_ELEC_URBAN, SPE_END_YEAR, SPE_MID_YEAR,
                       SPE_GRID_CAPACITY_INVESTMENT, SPE_GRID_LOSSES,
                       SPE_MAX_GRID_EXTENSION_DIST,
                       SPE_NUM_PEOPLE_PER_HH_RURAL,
                       SPE_NUM_PEOPLE_PER_HH_URBAN, SPE_POP, SPE_POP_FUTURE,
                       SPE_START_YEAR, SPE_URBAN, SPE_URBAN_FUTURE,
                       SPE_URBAN_MODELLED)
from openpyxl import load_workbook
import numpy as np

logging.basicConfig(format='%(asctime)s\t\t%(message)s', level=logging.DEBUG)


def calibration(specs_path, csv_path, specs_path_calib, calibrated_csv_path, scen):
    """

    Arguments
    ---------
    specs_path
    csv_path
    specs_path_calib
    calibrated_csv_path
    scen
    """
    specs_data = pd.read_excel(specs_path, sheet_name='SpecsData')
    specs_data_calib = pd.DataFrame()

    try:
        specs_data_calib = pd.read_excel(specs_path_calib, sheet_name='SpecsDataCalib')
    except:
        None

    settlements_in_csv = csv_path
    settlements_out_csv = calibrated_csv_path

    onsseter = SettlementProcessor(settlements_in_csv)

    num_people_per_hh_rural = float(specs_data.iloc[scen][SPE_NUM_PEOPLE_PER_HH_RURAL])
    num_people_per_hh_urban = float(specs_data.iloc[scen][SPE_NUM_PEOPLE_PER_HH_URBAN])
    ppl_hh_average = (num_people_per_hh_urban + num_people_per_hh_rural) / 2
    # RUN_PARAM: these are the annual household electricity targets
    tier_1 = ppl_hh_average * 8  # 38.7 refers to kWh/household/year. It is the mean value between Tier 1 and Tier 2
    tier_2 = ppl_hh_average * 44
    tier_3 = ppl_hh_average * 160
    tier_4 = ppl_hh_average * 423
    tier_5 = ppl_hh_average * 598

    onsseter.prepare_wtf_tier_columns(num_people_per_hh_rural, num_people_per_hh_urban,
                                      tier_1, tier_2, tier_3, tier_4, tier_5)
    onsseter.condition_df()
    onsseter.df[SET_GRID_PENALTY] = onsseter.grid_penalties(onsseter.df)

    onsseter.df[SET_WINDCF] = onsseter.calc_wind_cfs(onsseter.df[SET_WINDVEL])

    pop_actual = specs_data.iloc[scen][SPE_POP]
    urban_current = specs_data.iloc[scen][SPE_URBAN]
    start_year = int(specs_data.iloc[scen][SPE_START_YEAR])
    elec_actual = specs_data.iloc[scen][SPE_ELEC]
    elec_actual_urban = specs_data.iloc[scen][SPE_ELEC_URBAN]
    elec_actual_rural = specs_data.iloc[scen][SPE_ELEC_RURAL]

    if not specs_data_calib.empty:
        specs_data[SPE_ELEC_MODELLED] = specs_data_calib[SPE_ELEC_MODELLED]
        specs_data['rural_elec_ratio_modelled'] = specs_data_calib['rural_elec_ratio_modelled']
        specs_data['urban_elec_ratio_modelled'] = specs_data_calib['urban_elec_ratio_modelled']
        specs_data['grid_data_used'] = specs_data_calib['grid_data_used']
        specs_data['grid_distance_used'] = specs_data_calib['grid_distance_used']
        specs_data['ntl_limit'] = specs_data_calib['ntl_limit']
        specs_data['pop_limit'] = specs_data_calib['pop_limit']
        specs_data['Buffer_used'] = specs_data_calib['Buffer_used']
        specs_data['buffer_distance'] = specs_data_calib['buffer_distance']

    pop_modelled, urban_modelled = onsseter.calibrate_current_pop_and_urban(pop_actual, urban_current, scen)

    specs_data.iloc[scen][SPE_URBAN_MODELLED] = urban_modelled

    elec_calibration_results = onsseter.calibrate_elec_current(elec_actual, elec_actual_urban, elec_actual_rural,
                                                               start_year, scen, buffer=True)

    specs_data.loc[scen, SPE_ELEC_MODELLED] = elec_calibration_results[0]
    specs_data.loc[scen, 'rural_elec_ratio_modelled'] = elec_calibration_results[1]
    specs_data.loc[scen, 'urban_elec_ratio_modelled'] = elec_calibration_results[2]
    specs_data.loc[scen, 'grid_data_used'] = elec_calibration_results[3]
    specs_data.loc[scen, 'grid_distance_used'] = elec_calibration_results[4]
    specs_data.loc[scen, 'ntl_limit'] = elec_calibration_results[5]
    specs_data.loc[scen, 'pop_limit'] = elec_calibration_results[6]
    specs_data.loc[scen, 'Buffer_used'] = elec_calibration_results[7]
    specs_data.loc[scen, 'buffer_distance'] = elec_calibration_results[8]

    book = load_workbook(specs_path)
    writer = pd.ExcelWriter(specs_path_calib, engine='openpyxl')
    writer.book = book
    # RUN_PARAM: Here the calibrated "specs" data are copied to a new tab called "SpecsDataCalib". 
    # This is what will later on be used to feed the model
    specs_data.to_excel(writer, sheet_name='SpecsDataCalib', index=False)
    #writer.save()
    writer.close()

    logging.info('Calibration finished. Results are transferred to the csv file \n')
    onsseter.df.to_csv(settlements_out_csv, index=False)


def scenario(specs_path, calibrated_csv_path, results_folder, summary_folder, scen, specs_path_orig, csv_path):
    """

    Arguments
    ---------
    specs_path : str
    calibrated_csv_path : str
    results_folder : str
    summary_folder : str
    scen : int
    specs_path_orig : str
    csv_path : list[str]

    """

    scenario_info = pd.read_excel(specs_path, sheet_name='ScenarioInfo')
    # scenarios = scenario_info['Scenario']
    scenario_parameters = pd.read_excel(specs_path, sheet_name='ScenarioParameters')
    specs_data = pd.read_excel(specs_path, sheet_name='SpecsDataCalib')

    print(specs_data.loc[0, SPE_COUNTRY], 'Scenario: ' + str(scen + 1))
    country_id = specs_data.iloc[scen]['CountryCode']
    pop_future = specs_data.loc[scen, SPE_POP_FUTURE]
    urban_future = specs_data.loc[scen, SPE_URBAN_FUTURE]
    grid_base_to_peak = specs_data.iloc[scen]['BaseToPeak']

    pop_index = scenario_info.iloc[scen]['Population_Growth']
    tier_index = scenario_info.iloc[scen]['Target_electricity_consumption_level']
    mid_year_index = scenario_info.iloc[scen]['Electrification_target_mid_year']
    end_year_index = scenario_info.iloc[scen]['Electrification_target_end_year']
    grid_index = scenario_info.iloc[scen]['Grid_electricity_generation_cost']
    pv_index = scenario_info.iloc[scen]['PV_cost_adjust']
    diesel_index = scenario_info.iloc[scen]['Diesel_price']
    productive_index = scenario_info.iloc[scen]['Productive_uses_demand']
    prio_index = scenario_info.iloc[scen]['Prioritization_algorithm']
    discount_rate_index = scenario_info.iloc[scen]['Discount_rate']
    grid_om_index = scenario_info.iloc[scen]['Grid_OM_TDLines']
    grid_cxn_cost_index = scenario_info.iloc[scen]['Grid_connection_cost_per_HH']
    grid_cap_fac_index = scenario_info.iloc[scen]['Grid_capacity_factor']
    grid_tech_life_index = scenario_info.iloc[scen]['Grid_technical_life']
    grid_pnlty_index = scenario_info.iloc[scen]['Grid_penalty_ratio']
    mghydro_om_td_index = scenario_info.iloc[scen]['MGHydro_OM_TDLines']
    mghydro_dist_loss_index = scenario_info.iloc[scen]['MGHydro_distribution_losses']
    mghydro_cxn_cost_index = scenario_info.iloc[scen]['MGHydro_connection_cost_per_HH']
    mghydro_base_to_peak_index = scenario_info.iloc[scen]['MGHydro_base_to_peak_load_ratio']
    mghydro_cap_cost_index = scenario_info.iloc[scen]['MGHydro_capital_cost']
    mghydro_cap_fac_index = scenario_info.iloc[scen]['MGHydro_capacity_factor']
    mghydro_om_cost_index = scenario_info.iloc[scen]['MGHydro_OM_costs']
    mghydro_tech_life_index = scenario_info.iloc[scen]['MGHydro_technical_life']
    mgwind_om_td_index = scenario_info.iloc[scen]['MGWind_OM_TDLines']
    mgwind_dist_loss_index = scenario_info.iloc[scen]['MGWind_distribution_losses']
    mgwind_cxn_cost_index = scenario_info.iloc[scen]['MGWind_connection_cost_per_HH']
    mgwind_base_to_peak_index = scenario_info.iloc[scen]['MGWind_base_to_peak_load_ratio']
    mgwind_cap_cost_index = scenario_info.iloc[scen]['MGWind_capital_cost']
    mgwind_om_cost_index = scenario_info.iloc[scen]['MGWind_OM_costs']
    mgwind_tech_life_index = scenario_info.iloc[scen]['MGWind_technical_life']
    mgpv_om_td_index = scenario_info.iloc[scen]['MGPV_OM_TDLines']
    mgpv_dist_loss_index = scenario_info.iloc[scen]['MGPV_distribution_losses']
    mgpv_cxn_cost_index = scenario_info.iloc[scen]['MGPV_connection_cost_per_HH']
    mgpv_base_to_peak_index = scenario_info.iloc[scen]['MGPV_base_to_peak_load_ratio']
    mgpv_cap_cost_index = scenario_info.iloc[scen]['MGPV_capital_cost']
    mgpv_om_cost_index = scenario_info.iloc[scen]['MGPV_OM_costs']
    mgpv_tech_life_index = scenario_info.iloc[scen]['MGPV_technical_life']
    sapv_base_to_peak_index = scenario_info.iloc[scen]['SAPV_base_to_peak_load_ratio']
    sapv_cap_cost_inf_index = scenario_info.iloc[scen]['SAPV_capital_cost_Inf']
    sapv_cap_cost_1kw_index = scenario_info.iloc[scen]['SAPV_capital_cost_1kW']
    sapv_cap_cost_100w_index = scenario_info.iloc[scen]['SAPV_capital_cost_100W']
    sapv_cap_cost_50w_index = scenario_info.iloc[scen]['SAPV_capital_cost_50W']
    sapv_cap_cost_20w_index = scenario_info.iloc[scen]['SAPV_capital_cost_20W']
    sapv_om_cost_index = scenario_info.iloc[scen]['SAPV_OM_costs']
    sapv_tech_life_index = scenario_info.iloc[scen]['SAPV_technical_life']
    mgdiesel_om_td_index = scenario_info.iloc[scen]['MGDiesel_OM_TDLines']
    mgdiesel_dist_loss_index = scenario_info.iloc[scen]['MGDiesel_distribution_losses']
    mgdiesel_cxn_cost_index = scenario_info.iloc[scen]['MGDiesel_connection_cost_per_HH']
    mgdiesel_base_to_peak_index = scenario_info.iloc[scen]['MGDiesel_base_to_peak_load_ratio']
    mgdiesel_cap_cost_index = scenario_info.iloc[scen]['MGDiesel_capital_cost']
    mgdiesel_om_cost_index = scenario_info.iloc[scen]['MGDiesel_OM_costs']
    mgdiesel_tech_life_index = scenario_info.iloc[scen]['MGDiesel_technical_life']
    mgdiesel_cap_fac_index = scenario_info.iloc[scen]['MGDiesel_capacity_factor']
    sadiesel_base_to_peak_index = scenario_info.iloc[scen]['SADiesel_base_to_peak_load_ratio']
    sadiesel_cap_cost_index = scenario_info.iloc[scen]['SADiesel_capital_cost']
    sadiesel_om_cost_index = scenario_info.iloc[scen]['SADiesel_OM_costs']
    sadiesel_tech_life_index = scenario_info.iloc[scen]['SADiesel_technical_life']
    sadiesel_cap_fac_index = scenario_info.iloc[scen]['SADiesel_capacity_factor']
    sadiesel_eff_index = scenario_info.iloc[scen]['SADiesel_efficiency']
    sadiesel_truck_consump_index = scenario_info.iloc[scen]['SADiesel_truck_consumption']
    sadiesel_truck_vol_index = scenario_info.iloc[scen]['SADiesel_truck_volume']
    mgdiesel_eff_index = scenario_info.iloc[scen]['MGDiesel_efficiency']
    mgdiesel_truck_consump_index = scenario_info.iloc[scen]['MGDiesel_truck_consumption']
    mgdiesel_truck_vol_index = scenario_info.iloc[scen]['MGDiesel_truck_volume']

    end_year_pop = scenario_parameters.iloc[pop_index]['PopEndYear']
    rural_tier = scenario_parameters.iloc[tier_index]['RuralTargetTier']
    urban_tier = scenario_parameters.iloc[tier_index]['UrbanTargetTier']
    mid_year_target = scenario_parameters.iloc[mid_year_index]['MidYearTarget']
    end_year_target = scenario_parameters.iloc[end_year_index]['EndYearTarget']
    annual_new_grid_connections_limit = scenario_parameters.iloc[mid_year_index][
                                            'GridConnectionsLimitThousands'] * 1000
    grid_price = scenario_parameters.iloc[grid_index]['GridGenerationCost']
    pv_capital_cost_adjust = scenario_parameters.iloc[pv_index]['PV_Cost_adjust']
    diesel_price = scenario_parameters.iloc[diesel_index]['DieselPrice']
    productive_demand = scenario_parameters.iloc[productive_index]['ProductiveDemand']
    prioritization = scenario_parameters.iloc[prio_index]['PrioritizationAlgorithm']
    auto_intensification = scenario_parameters.iloc[prio_index]['AutoIntensificationKM']
    disc_rate = scenario_parameters.iloc[discount_rate_index]['DiscountRate']
    grid_om_td_lines = scenario_parameters.iloc[grid_om_index]['GridOMofTDLines']
    grid_cxn_cost_per_HH = scenario_parameters.iloc[grid_cxn_cost_index]['GridConnectionCostPerHH']
    grid_cap_fac = scenario_parameters.iloc[grid_cap_fac_index]['GridCapacityFactor']
    grid_tech_life = scenario_parameters.iloc[grid_tech_life_index]['GridTechnicalLife']
    grid_pnlty_ratio = scenario_parameters.iloc[grid_pnlty_index]['GridPenaltyRatio']
    mghydro_om_td_lines = scenario_parameters.iloc[mghydro_om_td_index]['MGHydroOMofTDLines']
    mghydro_dist_losses = scenario_parameters.iloc[mghydro_dist_loss_index]['MGHydroDistributionLosses']
    mghydro_cxn_cost_per_HH = scenario_parameters.iloc[mghydro_cxn_cost_index]['MGHydroConnectionCostPerHH']
    mghydro_base_to_peak = scenario_parameters.iloc[mghydro_base_to_peak_index]['MGHydroBaseToPeakLoadRatio']
    mghydro_cap_fac = scenario_parameters.iloc[mghydro_cap_fac_index]['MGHydroCapacityFactor']
    mghydro_cap_cost = scenario_parameters.iloc[mghydro_cap_cost_index]['MGHydroCapitalCost']
    mghydro_om_costs = scenario_parameters.iloc[mghydro_om_cost_index]['MGHydroOMCosts']
    mghydro_tech_life = scenario_parameters.iloc[mghydro_tech_life_index]['MGHydroTechnicalLife']
    mgwind_om_td_lines = scenario_parameters.iloc[mgwind_om_td_index]['MGWindOMofTDLines']
    mgwind_dist_losses = scenario_parameters.iloc[mgwind_dist_loss_index]['MGWindDistributionLosses']
    mgwind_cxn_cost_per_HH = scenario_parameters.iloc[mgwind_cxn_cost_index]['MGWindConnectionCostPerHH']
    mgwind_base_to_peak = scenario_parameters.iloc[mgwind_base_to_peak_index]['MGWindBaseToPeakLoadRatio']
    mgwind_cap_cost = scenario_parameters.iloc[mgwind_cap_cost_index]['MGWindCapitalCost']
    mgwind_om_costs = scenario_parameters.iloc[mgwind_om_cost_index]['MGWindOMCosts']
    mgwind_tech_life = scenario_parameters.iloc[mgwind_tech_life_index]['MGWindTechnicalLife']
    mgpv_om_td_lines = scenario_parameters.iloc[mgpv_om_td_index]['MGPVOMofTDLines']
    mgpv_dist_losses = scenario_parameters.iloc[mgpv_dist_loss_index]['MGPVDistributionLosses']
    mgpv_cxn_cost_per_HH = scenario_parameters.iloc[mgpv_cxn_cost_index]['MGPVConnectionCostPerHH']
    mgpv_base_to_peak = scenario_parameters.iloc[mgpv_base_to_peak_index]['MGPVBaseToPeakLoadRatio']
    mgpv_cap_cost = scenario_parameters.iloc[mgpv_cap_cost_index]['MGPVCapitalCost']
    mgpv_om_costs = scenario_parameters.iloc[mgpv_om_cost_index]['MGPVOMCosts']
    mgpv_tech_life = scenario_parameters.iloc[mgpv_tech_life_index]['MGPVTechnicalLife']
    sapv_base_to_peak = scenario_parameters.iloc[sapv_base_to_peak_index]['SAPVBaseToPeakLoadRatio']
    sapv_cap_cost_inf = scenario_parameters.iloc[sapv_cap_cost_inf_index]['SAPVCapitalCostInf']
    sapv_cap_cost_1kw = scenario_parameters.iloc[sapv_cap_cost_1kw_index]['SAPVCapitalCost1kW']
    sapv_cap_cost_100w = scenario_parameters.iloc[sapv_cap_cost_100w_index]['SAPVCapitalCost100W']
    sapv_cap_cost_50w = scenario_parameters.iloc[sapv_cap_cost_50w_index]['SAPVCapitalCost50W']
    sapv_cap_cost_20w = scenario_parameters.iloc[sapv_cap_cost_20w_index]['SAPVCapitalCost20W']
    sapv_om_costs = scenario_parameters.iloc[sapv_om_cost_index]['SAPVOMCosts']
    sapv_tech_life = scenario_parameters.iloc[sapv_tech_life_index]['SAPVTechnicalLife']
    mgdiesel_om_td_lines = scenario_parameters.iloc[mgdiesel_om_td_index]['MGDieselOMofTDLines']
    mgdiesel_dist_losses = scenario_parameters.iloc[mgdiesel_dist_loss_index]['MGDieselDistributionLosses']
    mgdiesel_cxn_cost_per_HH = scenario_parameters.iloc[mgdiesel_cxn_cost_index]['MGDieselConnectionCostPerHH']
    mgdiesel_base_to_peak = scenario_parameters.iloc[mgdiesel_base_to_peak_index]['MGDieselBaseToPeakLoadRatio']
    mgdiesel_cap_cost = scenario_parameters.iloc[mgdiesel_cap_cost_index]['MGDieselCapitalCost']
    mgdiesel_om_costs = scenario_parameters.iloc[mgdiesel_om_cost_index]['MGDieselOMCosts']
    mgdiesel_tech_life = scenario_parameters.iloc[mgdiesel_tech_life_index]['MGDieselTechnicalLife']
    mgdiesel_cap_fac = scenario_parameters.iloc[mgdiesel_cap_fac_index]['MGDieselCapacityFactor']
    sadiesel_base_to_peak = scenario_parameters.iloc[sadiesel_base_to_peak_index]['SADieselBaseToPeakLoadRatio']
    sadiesel_cap_cost = scenario_parameters.iloc[sadiesel_cap_cost_index]['SADieselCapitalCost']
    sadiesel_om_costs = scenario_parameters.iloc[sadiesel_om_cost_index]['SADieselOMCosts']
    sadiesel_tech_life = scenario_parameters.iloc[sadiesel_tech_life_index]['SADieselTechnicalLife']
    sadiesel_cap_fac = scenario_parameters.iloc[sadiesel_cap_fac_index]['SADieselCapacityFactor']
    sadiesel_efficiency = scenario_parameters.iloc[sadiesel_eff_index]['SADieselEfficiency']
    sadiesel_truck_consumption = scenario_parameters.iloc[sadiesel_truck_consump_index]['SADieselTruckConsumption']
    sadiesel_truck_volume = scenario_parameters.iloc[sadiesel_truck_vol_index]['SADieselTruckVolume']
    mgdiesel_efficiency = scenario_parameters.iloc[mgdiesel_eff_index]['MGDieselEfficiency']
    mgdiesel_truck_consumption = scenario_parameters.iloc[mgdiesel_truck_consump_index]['MGDieselTruckConsumption']
    mgdiesel_truck_volume = scenario_parameters.iloc[mgdiesel_truck_vol_index]['MGDieselTruckVolume']

    start_year = specs_data.iloc[scen][SPE_START_YEAR].astype('int64')
    mid_year = specs_data.iloc[scen][SPE_MID_YEAR].astype('int64')
    end_year = specs_data.iloc[scen][SPE_END_YEAR].astype('int64')

    settlements_in_csv = calibrated_csv_path
    settlements_out_csv = os.path.join(results_folder,
                                       '{}-1-{}_{}_{}_{}_{}_{}_{}-{}.csv'.format(country_id, pop_index, tier_index,
                                                                           mid_year_index, grid_index, pv_index,
                                                                           prio_index, start_year, end_year))
    summary_csv = os.path.join(summary_folder,
                               '{}-1-{}_{}_{}_{}_{}_{}_{}-{}_summary.csv'.format(country_id, pop_index, tier_index,
                                                                           mid_year_index, grid_index, pv_index,
                                                                           prio_index, start_year, end_year))

    onsseter = SettlementProcessor(settlements_in_csv)

    num_people_per_hh_rural = float(specs_data.iloc[scen][SPE_NUM_PEOPLE_PER_HH_RURAL])
    num_people_per_hh_urban = float(specs_data.iloc[scen][SPE_NUM_PEOPLE_PER_HH_URBAN])
    max_grid_extension_dist = float(specs_data.iloc[scen][SPE_MAX_GRID_EXTENSION_DIST])
    annual_grid_cap_gen_limit = specs_data.loc[scen, 'NewGridGenerationCapacityAnnualLimitMW'] * 1000

    # RUN_PARAM: Fill in general and technology specific parameters (e.g. discount rate, losses etc.)
    Technology.set_default_values(base_year=start_year,
                                  start_year=start_year,
                                  end_year=end_year,
                                  discount_rate=disc_rate)

    grid_calc = Technology(om_of_td_lines=grid_om_td_lines,
                           distribution_losses=float(specs_data.iloc[scen][SPE_GRID_LOSSES]),
                           connection_cost_per_hh=grid_cxn_cost_per_HH,
                           base_to_peak_load_ratio=grid_base_to_peak,
                           capacity_factor=grid_cap_fac,
                           tech_life=grid_tech_life,
                           grid_capacity_investment=float(specs_data.iloc[scen][SPE_GRID_CAPACITY_INVESTMENT]),
                           grid_penalty_ratio=grid_pnlty_ratio,
                           grid_price=grid_price)

    mg_hydro_calc = Technology(om_of_td_lines=mghydro_om_td_lines,
                               distribution_losses=mghydro_dist_losses,
                               connection_cost_per_hh=mghydro_cxn_cost_per_HH,
                               base_to_peak_load_ratio=mghydro_base_to_peak,
                               capacity_factor=mghydro_cap_fac,
                               tech_life=mghydro_tech_life,
                               capital_cost={float("inf"): mghydro_cap_cost},
                               om_costs=mghydro_om_costs,
                               mini_grid=True)

    mg_wind_calc = Technology(om_of_td_lines=mgwind_om_td_lines,
                              distribution_losses=mgwind_dist_losses,
                              connection_cost_per_hh=mgwind_cxn_cost_per_HH,
                              base_to_peak_load_ratio=mgwind_base_to_peak,
                              capital_cost={float("inf"): mgwind_cap_cost},
                              om_costs=mgwind_om_costs,
                              tech_life=mgwind_tech_life,
                              mini_grid=True)

    mg_pv_calc = Technology(om_of_td_lines=mgpv_om_td_lines,
                            distribution_losses=mgpv_dist_losses,
                            connection_cost_per_hh=mgpv_cxn_cost_per_HH,
                            base_to_peak_load_ratio=mgpv_base_to_peak,
                            capital_cost={float("inf"): mgpv_cap_cost},
                            om_costs=mgpv_om_costs,
                            tech_life=mgpv_tech_life,
                            mini_grid=True)

    sa_pv_calc = Technology(base_to_peak_load_ratio=sapv_base_to_peak,
                            tech_life=sapv_tech_life,
                            om_costs=sapv_om_costs,
                            capital_cost={float("inf"): sapv_cap_cost_inf * pv_capital_cost_adjust,
                                          1: sapv_cap_cost_1kw * pv_capital_cost_adjust,
                                          0.100: sapv_cap_cost_100w * pv_capital_cost_adjust,
                                          0.050: sapv_cap_cost_50w * pv_capital_cost_adjust,
                                          0.020: sapv_cap_cost_20w * pv_capital_cost_adjust
                                          },
                            standalone=True)

    mg_diesel_calc = Technology(om_of_td_lines=mgdiesel_om_td_lines,
                                distribution_losses=mgdiesel_dist_losses,
                                connection_cost_per_hh=mgdiesel_cxn_cost_per_HH,
                                base_to_peak_load_ratio=mgdiesel_base_to_peak,
                                capacity_factor=mgdiesel_cap_fac,
                                tech_life=mgdiesel_tech_life,
                                om_costs=mgdiesel_om_costs,
                                capital_cost={float("inf"): mgdiesel_cap_cost},
                                mini_grid=True)

    sa_diesel_calc = Technology(base_to_peak_load_ratio=sadiesel_base_to_peak,
                                capacity_factor=sadiesel_cap_fac,
                                tech_life=sadiesel_tech_life,
                                om_costs=sadiesel_om_costs,
                                capital_cost={float("inf"): sadiesel_cap_cost},
                                standalone=True)

    sa_diesel_cost = {'diesel_price': diesel_price,
                      'efficiency': sadiesel_efficiency,
                      'diesel_truck_consumption': sadiesel_truck_consumption,
                      'diesel_truck_volume': sadiesel_truck_volume}

    mg_diesel_cost = {'diesel_price': diesel_price,
                      'efficiency': mgdiesel_efficiency,
                      'diesel_truck_consumption': mgdiesel_truck_consumption,
                      'diesel_truck_volume': mgdiesel_truck_volume}

    # RUN_PARAM: One shall define here the years of analysis (excluding start year),
    # together with access targets per interval and timestep duration
    yearsofanalysis = [mid_year, end_year]
    eleclimits = {mid_year: mid_year_target, end_year: end_year_target}
    time_steps = {mid_year: mid_year-start_year, end_year: end_year-mid_year}

    elements = ["1.Population", "2.New_Connections", "3.Capacity", "4.Investment"]
    techs = ["Grid", "SA_Diesel", "SA_PV", "MG_Diesel", "MG_PV", "MG_Wind", "MG_Hydro"]
    tech_codes = [1, 2, 3, 4, 5, 6, 7]

    sumtechs = []
    for element in elements:
        for tech in techs:
            sumtechs.append(element + "_" + tech)
    total_rows = len(sumtechs)
    df_summary = pd.DataFrame(columns=yearsofanalysis)
    for row in range(0, total_rows):
        df_summary.loc[sumtechs[row]] = "Nan"

    onsseter.current_mv_line_dist()

    onsseter.project_pop_and_urban(pop_future, urban_future, start_year, yearsofanalysis)

    for year in yearsofanalysis:
        eleclimit = eleclimits[year]
        time_step = time_steps[year]

        if year - time_step == start_year:
            grid_cap_gen_limit = time_step * annual_grid_cap_gen_limit
            grid_connect_limit = time_step * annual_new_grid_connections_limit
        else:
            grid_cap_gen_limit = 9999999999
            grid_connect_limit = 9999999999

        onsseter.set_scenario_variables(year, num_people_per_hh_rural, num_people_per_hh_urban, time_step,
                                        start_year, urban_tier, rural_tier, end_year_pop, productive_demand)

        onsseter.diesel_cost_columns(sa_diesel_cost, mg_diesel_cost, year)

        sa_diesel_investment, sa_diesel_capacity, sa_pv_investment, sa_pv_capacity, mg_diesel_investment, \
        mg_diesel_capacity, mg_pv_investment, mg_pv_capacity, mg_wind_investment, mg_wind_capacity, \
        mg_hydro_investment, mg_hydro_capacity = onsseter.calculate_off_grid_lcoes(mg_hydro_calc, mg_wind_calc, mg_pv_calc,
                                                                    sa_pv_calc, mg_diesel_calc,
                                                                    sa_diesel_calc, year, end_year, time_step,
                                                                    techs, tech_codes)

        grid_investment, grid_capacity, grid_cap_gen_limit, grid_connect_limit = \
            onsseter.pre_electrification(grid_price, year, time_step, end_year, grid_calc, grid_cap_gen_limit,
                                         grid_connect_limit)

        onsseter.df[SET_LCOE_GRID + "{}".format(year)], onsseter.df[SET_MIN_GRID_DIST + "{}".format(year)], \
        onsseter.df[SET_ELEC_ORDER + "{}".format(year)], onsseter.df[SET_MV_CONNECT_DIST], grid_investment,\
            grid_capacity = \
            onsseter.elec_extension(grid_calc,
                                    max_grid_extension_dist,
                                    year,
                                    start_year,
                                    end_year,
                                    time_step,
                                    grid_cap_gen_limit,
                                    grid_connect_limit,
                                    auto_intensification=auto_intensification,
                                    prioritization=prioritization,
                                    new_investment=grid_investment,
                                    new_capacity=grid_capacity)

        onsseter.results_columns(techs, tech_codes, year, time_step, prioritization, auto_intensification)

        onsseter.calculate_investments_and_capacity(sa_diesel_investment, sa_diesel_capacity, sa_pv_investment,
                                                    sa_pv_capacity, mg_diesel_investment, mg_diesel_capacity,
                                                    mg_pv_investment, mg_pv_capacity, mg_wind_investment,
                                                    mg_wind_capacity, mg_hydro_investment, mg_hydro_capacity,
                                                    grid_investment, grid_capacity, year)

        onsseter.apply_limitations(eleclimit, year, time_step, prioritization, auto_intensification)

        onsseter.calc_summaries(df_summary, sumtechs, tech_codes, year)

    for i in range(len(onsseter.df.columns)):
        if onsseter.df.iloc[:, i].dtype == 'float64':
            onsseter.df.iloc[:, i] = pd.to_numeric(onsseter.df.iloc[:, i], downcast='float')
        elif onsseter.df.iloc[:, i].dtype == 'int64':
            onsseter.df.iloc[:, i] = pd.to_numeric(onsseter.df.iloc[:, i], downcast='signed')

    df_summary.to_csv(summary_csv, index=sumtechs)
    onsseter.df.to_csv(settlements_out_csv, index=False)

    logging.info('Finished \n')

    return update(onsseter.df, scen, specs_path_orig, [start_year, mid_year, end_year], csv_path)


def update(results_df, scen, specs_path_orig, analysis_years, csv_path):
    """

    Arguments
    ---------
    results_df : Dataframe
    scen : int
    specs_path_orig : str
    analysis_years : list[int]
    csv_path : list[str]

    """

    new_grid_demand = []
    elecstatusin_end_year = "ElecStatusIn" + str(analysis_years[2])
    pop_end_year = 'Pop'+str(analysis_years[2])
    finaleleccode_end_year = "FinalElecCode"+str(analysis_years[2])

    # Update GIS file for the subsequent scenario run

    specs_data = pd.read_excel(specs_path_orig, sheet_name='SpecsData')
    try:
        urban_future = specs_data["UrbanRatioStartYear"][scen + 1]
    except:
        urban_future = specs_data["UrbanRatioStartYear"][scen] #not ever used, exists so that the code doesnt break

    results_df.sort_values(by=[pop_end_year], inplace=True, ascending=False)
    cumulative_pop = results_df[pop_end_year].cumsum()
    results_df['IsUrban'] = np.where(cumulative_pop < (urban_future * results_df[pop_end_year].sum()), 2, 0)
    results_df.sort_index(inplace=True)

    gis_data_future = pd.read_csv(csv_path[scen])
    gis_data_future['Pop'] = results_df[pop_end_year]
    gis_data_future['IsUrban'] = results_df['IsUrban']
    gis_data_future['ElecPop'] = results_df[pop_end_year] * results_df[elecstatusin_end_year]
    gis_data_future['ElectrificationOrder'] = results_df['ElectrificationOrder'+str(analysis_years[2])]
    gis_data_future["ElecStart"] = results_df[elecstatusin_end_year]
    gis_data_future[finaleleccode_end_year] = results_df[finaleleccode_end_year]

    try:
        future_csv_path = csv_path[scen+1]
        gis_data_future.to_csv(future_csv_path, index=False)
    except:
        pass

    # Update electrification ratio in specs file

    rural_elec_pop_end_year = gis_data_future.loc[gis_data_future["IsUrban"] < 2, 'ElecPop'].sum()
    urban_elec_pop_end_year = gis_data_future.loc[gis_data_future["IsUrban"] == 2, 'ElecPop'].sum()
    total_elec_pop_end_year = rural_elec_pop_end_year + urban_elec_pop_end_year

    rural_pop_end_year = results_df.loc[results_df["IsUrban"] < 2, pop_end_year].sum()
    urban_pop_end_year = results_df.loc[results_df["IsUrban"] == 2, pop_end_year].sum()
    total_pop_end_year = results_df[pop_end_year].sum()

    urban_pop_ratio_future = urban_pop_end_year/total_pop_end_year

    specs_data["ElecActual"][scen + 1] = total_elec_pop_end_year / total_pop_end_year
    specs_data["Urban_elec_ratio"][scen + 1] = urban_elec_pop_end_year / urban_pop_end_year
    specs_data["Rural_elec_ratio"][scen + 1] = rural_elec_pop_end_year / rural_pop_end_year
    # RUN_PARAM: comment the line below out if you want to use the urban pop ratios in the start year from the specs file
    # The below line derives the urban starting ratio from the end
    #specs_data["UrbanRatioStartYear"][scen + 1] = urban_pop_end_year / total_pop_end_year

    #specs_data["PopStartYear"][scen + 1] = results_df[pop_end_year].sum()

    book = load_workbook(specs_path_orig)
    writer = pd.ExcelWriter(specs_path_orig, engine='openpyxl')
    writer.book = book
    std = book.get_sheet_by_name('SpecsData')
    book.remove_sheet(std)
    specs_data.to_excel(writer, sheet_name='SpecsData', index=False)
    #writer.save()
    writer.close()

    # Collect new demand from OnSSET

    for y in range(1, len(analysis_years)):
        energypersettlement = 'EnergyPerSettlement' + str(analysis_years[y])
        finaleleccode = 'FinalElecCode' + str(analysis_years[y])

        annual_new_demand = results_df.loc[results_df[finaleleccode] == 1, energypersettlement].sum()/(analysis_years[y]-analysis_years[y-1])

        for yy in range(analysis_years[y-1], analysis_years[y]):
            new_grid_demand.append(annual_new_demand)

    return new_grid_demand








